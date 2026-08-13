"""
Agent 工具函数
- geocode: 高德地图地址核验
- web_search: 联网搜索
- parse_excel: Excel 文件解析
- ocr_image: 图片 OCR 文字提取
"""

import os
import json
import base64
import httpx
from io import BytesIO
from pathlib import Path
from typing import Any
from dotenv import load_dotenv

load_dotenv()
# 兜底：若用户未创建 .env，则从 .env.example 载入预配 Key（如高德），实现开箱即用
_ENV_EXAMPLE = Path(__file__).resolve().parent.parent / ".env.example"
if _ENV_EXAMPLE.exists():
    load_dotenv(_ENV_EXAMPLE, override=False)

# Streamlit Cloud: read secrets from st.secrets, fallback to env vars
def _get_config(key: str, default: str = "") -> str:
    try:
        import streamlit as st
        return st.secrets.get(key, os.getenv(key, default))
    except Exception:
        return os.getenv(key, default)

# ============================================================
# 工具定义（OpenAI function calling 格式，传给 LLM）
# ============================================================

TOOL_DEFINITIONS = [
    {
        "type": "function",
        "function": {
            "name": "geocode",
            "description": "高德地图地址核验：传入地址字符串，返回该地址在地图上的定位结果，包括标准化地址、经纬度、匹配精度级别（兴趣点/门牌号/道路/区县/城市/省份）。用于验证地址是否真实存在。",
            "parameters": {
                "type": "object",
                "properties": {
                    "address": {
                        "type": "string",
                        "description": "需要验证的完整中文地址，如'上海市浦东新区张江路498号'",
                    },
                },
                "required": ["address"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "web_search",
            "description": "联网搜索：传入搜索关键词，返回网页搜索结果（标题、URL、摘要）。用于在地图之外补充验证地址是否在公开信息中被提及。",
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {
                        "type": "string",
                        "description": "搜索关键词，如'北京市海淀区中关村大街1号 地址'",
                    },
                },
                "required": ["query"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "parse_excel",
            "description": "解析上传的 Excel/CSV 文件，提取其中的客户姓名和地址信息。支持 .xlsx 和 .csv 格式。",
            "parameters": {
                "type": "object",
                "properties": {
                    "file_base64": {
                        "type": "string",
                        "description": "Excel文件的 base64 编码内容",
                    },
                },
                "required": ["file_base64"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "ocr_image",
            "description": "OCR 图片文字识别：传入图片的 base64 编码，提取图片中的所有文字内容。用于识别截图、照片中的地址信息。返回提取到的文本内容。",
            "parameters": {
                "type": "object",
                "properties": {
                    "file_base64": {
                        "type": "string",
                        "description": "图片文件的 base64 编码字符串（支持 png/jpg/jpeg 格式）",
                    },
                },
                "required": ["file_base64"],
            },
        },
    },
]


# ============================================================
# 工具实现
# ============================================================


# 简单缓存：避免批量文件中同地址重复调 API
_geocode_cache: dict[str, str] = {}

def geocode(address: str) -> str:
    """
    高德地图地理编码 API
    文档：https://lbs.amap.com/api/webservice/guide/api/georegeo

    返回 JSON 字符串，包含：
    - status: "success" | "error"
    - found: bool 是否找到
    - formatted_address: 标准化地址
    - location: 经纬度 "lng,lat"
    - level: 匹配级别（兴趣点/门牌号/道路/区县/城市/省份）
    - province, city, district, street, number: 结构化地址
    - source: "高德地图地理编码API"
    """
    # 缓存命中：批量文件中同地址不重复调 API
    cache_key = address.strip()
    if cache_key in _geocode_cache:
        return _geocode_cache[cache_key]

    api_key = _get_config("AMAP_API_KEY")
    if not api_key:
        return json.dumps(
            {"status": "error", "message": "高德地图 API Key 未配置，无法进行地图核验"},
            ensure_ascii=False,
        )

    try:
        # 高德地理编码：直接传完整地址字符串（无需单独提取 city 参数，
        # 高德服务端会自动解析省/市/区/街道/门牌号各级信息）
        url = "https://restapi.amap.com/v3/geocode/geo"
        params: dict[str, Any] = {
            "key": api_key,
            "address": address,
        }

        response = httpx.get(url, params=params, timeout=10.0)
        data = response.json()

        # 🔴-2：传输/服务层错误（Key 失效 / 配额耗尽 / QPS 超限）必须上报 error，
        # 绝不能伪装成 found=False（那会被误判为"地址不存在" → 整表误判无效且无护栏拦截）
        if data.get("status") != "1":
            return json.dumps(
                {
                    "status": "error",
                    "message": (
                        f"高德地图 API 调用失败：{data.get('info', '未知错误')}"
                        f"（infocode={data.get('infocode', '?')}）。本条地址未完成地图核验，"
                        f"请勿据此判定该地址无效，应判为「不确定」。"
                    ),
                    "source": "高德地图地理编码API",
                },
                ensure_ascii=False,
            )

        geocodes = data.get("geocodes") or []
        if not geocodes:
            return json.dumps(
                {
                    "status": "success",
                    "found": False,
                    "message": f"高德地图未找到该地址：{address}",
                    "source": "高德地图地理编码API",
                },
                ensure_ascii=False,
            )

        # 🔴-3：暴露唯一性判定所需字段，避免模型误判"有效地址"
        # （标准二要求"只能匹配到一个地点"，高德返回 count>1 时应判"不确定"）
        match_count = int(data.get("count", len(geocodes)) or len(geocodes))
        geocode_info = geocodes[0]
        location = geocode_info.get("location", "")
        level = geocode_info.get("level", "")

        # 生成高德地图可跳转链接
        #  - map_url：搜索直达页（按地址名搜索）
        #  - marker_url：精确定位到坐标的 marker 页（供审核依据精确跳转）
        from urllib.parse import quote
        addr_query = geocode_info.get("formatted_address", address)
        map_url = f"https://ditu.amap.com/search?query={quote(addr_query)}" if location else ""
        marker_url = (
            f"https://uri.amap.com/marker?position={location}"
            f"&name={quote(addr_query)}&src=address-audit-agent"
            f"&coordinate=gaode&callnative=0"
            if location
            else ""
        )

        # 解析结构化地址信息
        result = {
            "status": "success",
            "found": True,
            "match_count": match_count,
            "is_unique": match_count == 1,
            "other_candidates": [g.get("formatted_address", "") for g in geocodes[1:4]],
            "formatted_address": geocode_info.get("formatted_address", address),
            "location": location,
            "map_url": map_url,
            "marker_url": marker_url,
            "level": level,
            "level_desc": _level_description(level),
            "province": geocode_info.get("province", ""),
            "city": geocode_info.get("city", ""),
            "district": geocode_info.get("district", ""),
            "street": geocode_info.get("street", ""),
            "number": geocode_info.get("number", ""),
            "source": "高德地图地理编码API",
        }
        json_result = json.dumps(result, ensure_ascii=False)
        _geocode_cache[cache_key] = json_result
        return json_result

    except Exception as e:
        return json.dumps(
            {"status": "error", "message": f"高德地图 API 调用失败：{str(e)}"},
            ensure_ascii=False,
        )


def _level_description(level: str) -> str:
    """将高德返回的 level 代码转为中文说明（对齐高德官方枚举，避免"未知精度级别"误导模型）"""
    mapping = {
        "兴趣点": "精确匹配到具体地点（最高精度）",
        "门址": "精确匹配到门牌号（最高精度之一）",
        "单元号": "精确匹配到楼栋/单元",
        "道路交叉路口": "匹配到路口级别",
        "道路": "匹配到道路级别，具体门牌号可能不精确",
        "热点商圈": "匹配到商圈级别",
        "村庄": "仅匹配到村庄级，缺门牌号",
        "乡镇": "仅匹配到乡镇级",
        "区县": "仅匹配到区县级，地址可能不完整",
        "开发区": "仅匹配到开发区级",
        "城市": "仅匹配到城市级，地址信息严重不足",
        "省": "仅匹配到省级",
        "省份": "仅匹配到省级",
        "国家": "仅匹配到国家级",
    }
    return mapping.get(level, f"未知精度级别：{level}")


# 浏览器 UA（免费网页搜索源需要，否则易被拒）
_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/120.0 Safari/537.36"
)


def web_search(query: str) -> str:
    """
    联网搜索，多级回退（保证国内网络环境下不超时、有结果）：
      1. Bing Web Search API（若配置了 BING_SEARCH_KEY）
      2. Bing 国内版网页搜索（cn.bing.com，免费、国内直连、速度快）
      3. 百度搜索（免费、国内可达；结果为百度跳转链接，点击仍可直达原文）
      4. DuckDuckGo（免费，但国内通常不可达，仅作兜底）

    返回 JSON 字符串，包含：
    - status: "success" | "error"
    - results: [{title, url, snippet}, ...]
    - source: 实际命中的搜索源
    """
    api_key = _get_config("BING_SEARCH_KEY")
    if api_key and api_key != "your-bing-key":
        out = _bing_api_search(query, api_key)
        if out:
            return out

    # 免费回退链：取第一个「有结果」的源；都没有结果则返回第一个成功响应（空结果也是证据）
    first_success: str | None = None
    for fn in (_bing_cn_search, _baidu_search, _duckduckgo_search):
        try:
            out = fn(query)
            data = json.loads(out)
            if data.get("status") != "success":
                continue
            if first_success is None:
                first_success = out
            if data.get("total", 0) > 0:
                return out
        except Exception:
            continue

    if first_success is not None:
        return first_success
    return json.dumps(
        {"status": "error", "message": "联网搜索暂不可用（各搜索源均无响应）", "source": "web_search"},
        ensure_ascii=False,
    )


def _bing_api_search(query: str, api_key: str) -> str | None:
    """Bing Web Search API（需 Key）；失败返回 None 以便回退"""
    try:
        url = "https://api.bing.microsoft.com/v7.0/search"
        headers = {"Ocp-Apim-Subscription-Key": api_key}
        params: dict[str, Any] = {"q": query, "count": 5, "mkt": "zh-CN"}
        response = httpx.get(url, headers=headers, params=params, timeout=8.0)
        # 🟠-13：非 2xx（如 Key 失效返回 401/403）应触发回退链，而非返回一份「success 但 0 结果」
        # 的伪证据（否则会误导上层把"坏 Key"当成"无结果"处理）。
        if response.status_code != 200:
            return None
        data = response.json()
        results = [
            {"title": i.get("name", ""), "url": i.get("url", ""), "snippet": i.get("snippet", "")}
            for i in data.get("webPages", {}).get("value", [])
        ]
        return json.dumps(
            {"status": "success", "results": results, "total": len(results), "source": "Bing Web Search API"},
            ensure_ascii=False,
        )
    except Exception:
        return None


def _bing_cn_search(query: str) -> str:
    """Bing 国内版网页搜索（免费、无需 Key、国内直连，实测 ~0.6s）"""
    import re as _re
    from urllib.parse import quote as _quote

    resp = httpx.get(
        f"https://cn.bing.com/search?q={_quote(query)}",
        headers={"User-Agent": _UA},
        timeout=5.0,
        follow_redirects=True,
    )
    html = resp.text
    results = []
    for block in _re.findall(r'<li class="b_algo".*?</li>', html, _re.DOTALL):
        m = _re.search(r'<h2[^>]*><a[^>]*href="([^"]+)"[^>]*>(.*?)</a>', block, _re.DOTALL)
        if not m:
            continue
        url = m.group(1)
        title = _re.sub(r"<[^>]+>", "", m.group(2)).strip()
        snip_m = _re.search(r"<p[^>]*>(.*?)</p>", block, _re.DOTALL)
        snippet = _re.sub(r"<[^>]+>", "", snip_m.group(1)).strip() if snip_m else ""
        if url and title:
            results.append({"title": title, "url": url, "snippet": snippet})
        if len(results) >= 5:
            break
    return json.dumps(
        {"status": "success", "results": results, "total": len(results), "source": "Bing 国内版网页搜索"},
        ensure_ascii=False,
    )


def _baidu_search(query: str) -> str:
    """百度搜索（免费、国内可达；结果为百度跳转链接，点击仍可直达原文）"""
    import re as _re
    from urllib.parse import quote as _quote

    resp = httpx.get(
        f"https://www.baidu.com/s?wd={_quote(query)}",
        headers={"User-Agent": _UA},
        timeout=5.0,
        follow_redirects=True,
    )
    html = resp.text
    results = []
    for m in _re.finditer(
        r'<h3[^>]*class="[^"]*t[^"]*"[^>]*>\s*<a[^>]*href="([^"]+)"[^>]*>(.*?)</a>',
        html,
        _re.DOTALL,
    ):
        url = m.group(1)
        title = _re.sub(r"<[^>]+>", "", m.group(2)).strip()
        if url and title:
            results.append({"title": title, "url": url, "snippet": ""})
        if len(results) >= 5:
            break
    return json.dumps(
        {"status": "success", "results": results, "total": len(results), "source": "百度搜索"},
        ensure_ascii=False,
    )


def _duckduckgo_search(query: str) -> str:
    """DuckDuckGo 免费搜索（无需 Key；国内通常不可达，仅作最后兜底，超时 4s 快速失败）"""
    import re
    from html.parser import HTMLParser

    class DDGParser(HTMLParser):
        def __init__(self):
            super().__init__()
            self.results = []
            self.current = None

        def handle_starttag(self, tag, attrs):
            attrs = dict(attrs)
            if tag == "a" and "result-link" in attrs.get("class", ""):
                self.current = {"title": "", "url": attrs.get("href", ""), "snippet": ""}

        def handle_endtag(self, tag):
            if tag == "a" and self.current:
                self.results.append(self.current)
                self.current = None

        def handle_data(self, data):
            if self.current:
                self.current["title"] += data.strip()

    try:
        resp = httpx.get(
            "https://lite.duckduckgo.com/lite/",
            params={"q": query},
            headers={"User-Agent": _UA},
            timeout=4.0,
        )
        html = resp.text
        parser = DDGParser()
        parser.feed(html)

        for r in parser.results:
            if r["url"]:
                try:
                    pat = re.escape(r["url"]) + r'.*?<span class="result-snippet">(.*?)</span>'
                    m = re.search(pat, html, re.DOTALL)
                    if m:
                        r["snippet"] = re.sub(r"<[^>]+>", "", m.group(1)).strip()
                except re.error:
                    # URL 含正则特殊字符导致匹配失败，跳过 snippet 提取即可
                    pass

        results = parser.results[:5]
        return json.dumps(
            {"status": "success", "results": results, "total": len(results),
             "source": "DuckDuckGo (免费搜索)"},
            ensure_ascii=False,
        )
    except Exception as e:
        return json.dumps(
            {"status": "error", "message": f"搜索失败：{str(e)}", "source": "DuckDuckGo"},
            ensure_ascii=False,
        )


# 列名候选词表：Excel / CSV 共用同一份，避免两边解析口径不一致（🟠-17 修复）
_NAME_CANDIDATES = ["姓名", "客户名称", "公司名称", "name", "客户", "联系人", "UID", "客户UID"]
_ADDR_CANDIDATES = [
    "地址", "详细地址", "联系地址", "address", "addr", "公司地址", "注册地址", "客户地址"
]


def parse_excel(file_base64: str) -> str:
    """
    解析表格文件（Excel 或 CSV），提取姓名和地址列

    返回 JSON 字符串，包含：
    - status: "success" | "error"
    - records: [{name, address, row}, ...]
    - total: 记录总数
    - columns: 检测到的列名
    """
    # 解码 base64
    file_bytes = base64.b64decode(file_base64)

    # 先尝试 Excel 格式
    try:
        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(file_bytes), read_only=True)
    except Exception:
        # Excel 解析失败，尝试 CSV
        return _parse_csv(file_bytes)

    try:
        ws = wb.active

        if ws is None:
            return json.dumps(
                {"status": "error", "message": "Excel 文件中没有可读取的工作表"},
                ensure_ascii=False,
            )

        # 读取表头
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return json.dumps(
                {"status": "error", "message": "Excel 文件为空"},
                ensure_ascii=False,
            )

        headers = [str(h).strip() if h else "" for h in rows[0]]
        data_rows = rows[1:]

        # 智能匹配列名（支持中英文常见写法；与 _parse_csv 共用候选词表）
        name_col = _find_column(headers, _NAME_CANDIDATES)
        addr_col = _find_column(headers, _ADDR_CANDIDATES)

        if addr_col is None:
            return json.dumps(
                {
                    "status": "error",
                    "message": f"未找到地址列。检测到的列名：{headers}。请确保 Excel 中包含'地址'或'详细地址'列。",
                    "columns": headers,
                },
                ensure_ascii=False,
            )

        records = []
        for i, row in enumerate(data_rows):
            if not row:
                continue
            addr = str(row[addr_col]).strip() if len(row) > addr_col and row[addr_col] else ""
            if not addr or addr == "None":
                continue

            name = ""
            if name_col is not None and len(row) > name_col:
                name = str(row[name_col]).strip() if row[name_col] else ""

            records.append(
                {
                    "name": name if name and name != "None" else "未知",
                    "address": addr,
                    "row": i + 2,  # Excel 行号（从 1 开始，第 1 行是表头）
                }
            )

        return json.dumps(
            {
                "status": "success",
                "records": records,
                "total": len(records),
                "columns": headers,
                "name_column": headers[name_col] if name_col is not None else None,
                "address_column": headers[addr_col],
            },
            ensure_ascii=False,
        )

    except Exception as e:
        return json.dumps(
            {"status": "error", "message": f"Excel 解析失败：{str(e)}"},
            ensure_ascii=False,
        )
    finally:
        wb.close()


def _parse_csv(file_bytes: bytes) -> str:
    """CSV 文件解析，自动检测编码和分隔符"""
    import csv
    import io

    # 自动检测编码
    text = None
    for enc in ["utf-8-sig", "utf-8", "gbk", "gb2312", "latin-1"]:
        try:
            text = file_bytes.decode(enc)
            break
        except (UnicodeDecodeError, LookupError):
            continue

    if text is None:
        return json.dumps(
            {"status": "error", "message": "CSV 文件编码无法识别"},
            ensure_ascii=False,
        )

    # 快速检测分隔符：跳过慢速 csv.Sniffer，直接看第一行
    first_line = text.split("\n")[0].replace("\r", "") if text else ""
    delim = "\t" if first_line.count("\t") > first_line.count(",") else ","

    reader = csv.reader(io.StringIO(text), delimiter=delim)
    try:
        headers = [h.strip() if h else "" for h in next(reader)]
    except StopIteration:
        return json.dumps(
            {"status": "error", "message": "CSV 文件为空"},
            ensure_ascii=False,
        )

    # 使用与 Excel 解析相同的列匹配逻辑（共用候选词表）
    name_col = _find_column(headers, _NAME_CANDIDATES)
    addr_col = _find_column(headers, _ADDR_CANDIDATES)

    if addr_col is None:
        return json.dumps(
            {
                "status": "error",
                "message": f"未找到地址列。检测到的列名：{headers}。请确保 CSV 中包含'地址'或'详细地址'列。",
                "columns": headers,
            },
            ensure_ascii=False,
        )

    records = []
    for i, row in enumerate(reader, start=2):
        if not row or all(not cell for cell in row):
            continue
        addr = str(row[addr_col]).strip() if len(row) > addr_col and row[addr_col] else ""
        if not addr or addr == "None":
            continue
        name = ""
        if name_col is not None and len(row) > name_col and row[name_col]:
            name = str(row[name_col]).strip()

        records.append(
            {
                "name": name if name and name != "None" else "未知",
                "address": addr,
                "row": i,
            }
        )

    return json.dumps(
        {
            "status": "success",
            "records": records,
            "total": len(records),
            "columns": headers,
            "name_column": headers[name_col] if name_col is not None else None,
            "address_column": headers[addr_col],
        },
        ensure_ascii=False,
    )


def _find_column(headers: list[str], candidates: list[str]) -> int | None:
    """在表头中智能匹配目标列名（优先精确匹配，避免误匹配如'邮箱地址'→'地址'）"""
    for i, h in enumerate(headers):
        h_clean = h.lower().replace(" ", "").replace("_", "")
        for c in candidates:
            c_clean = c.lower().replace(" ", "").replace("_", "")
            # 精确匹配优先
            if h_clean == c_clean:
                return i
            # 长候选词（≥3字）允许子串匹配，短候选词仅精确匹配
            if len(c_clean) >= 3 and c_clean in h_clean:
                return i
    return None


def ocr_image(file_base64: str) -> str:
    """
    OCR 图片文字识别
    使用 pytesseract 提取图片中的文字，用于识别截图/照片中的地址信息。

    返回 JSON 字符串，包含：
    - status: "success" | "error"
    - text: 识别出的完整文本
    - source: "pytesseract OCR"
    """
    try:
        from PIL import Image

        img_bytes = base64.b64decode(file_base64)
        img = Image.open(BytesIO(img_bytes))
    except Exception as e:
        return json.dumps(
            {"status": "error", "message": f"图片解码失败：{str(e)}"},
            ensure_ascii=False,
        )

    try:
        import pytesseract

        text = pytesseract.image_to_string(img, lang="chi_sim+eng")
        text = text.strip()

        if not text:
            return json.dumps(
                {
                    "status": "success",
                    "text": "",
                    "message": "图片中未识别到文字，请确认图片清晰且包含文字内容。",
                    "source": "pytesseract OCR",
                },
                ensure_ascii=False,
            )

        return json.dumps(
            {
                "status": "success",
                "text": text,
                "source": "pytesseract OCR",
            },
            ensure_ascii=False,
        )
    except ImportError:
        return json.dumps(
            {
                "status": "error",
                "message": "OCR 功能未安装。请运行 pip install pytesseract Pillow 并安装 Tesseract-OCR。\n下载地址：https://github.com/UB-Mannheim/tesseract/wiki",
                "source": "pytesseract OCR（未安装）",
            },
            ensure_ascii=False,
        )
    except Exception as e:
        return json.dumps(
            {"status": "error", "message": f"OCR 识别失败：{str(e)}"},
            ensure_ascii=False,
        )


# ============================================================
# 工具执行调度
# ============================================================

TOOL_FUNCTIONS = {
    "geocode": geocode,
    "web_search": web_search,
    "parse_excel": parse_excel,
    "ocr_image": ocr_image,
}


def execute_tool(name: str, arguments: dict) -> str:
    """根据工具名称执行对应的工具函数"""
    import inspect

    func = TOOL_FUNCTIONS.get(name)
    if func is None:
        return json.dumps(
            {"status": "error", "message": f"未知工具：{name}"},
            ensure_ascii=False,
        )

    # 只传递函数签名中存在的参数，过滤 LLM 可能多余生成的 key
    valid_params = set(inspect.signature(func).parameters.keys())
    filtered_args = {k: v for k, v in arguments.items() if k in valid_params}

    try:
        return func(**filtered_args)
    except Exception as e:
        return json.dumps(
            {"status": "error", "message": f"工具 {name} 执行失败：{str(e)}"},
            ensure_ascii=False,
        )
